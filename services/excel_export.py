"""Excel export service — produces a professionally styled .xlsx workbook."""

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlmodel import Session, select

from i18n import _
from models.movement import Movement, MovementTag
from models.recurring import RecurringItem
from models.saving import Saving, SavingTag
from models.source import Source
from models.tag import Tag
from models.whim import Whim

# ---------------------------------------------------------------------------
# Colour palette (Sneat-inspired)
# ---------------------------------------------------------------------------
_PRIMARY = "696CFF"
_PRIMARY_LIGHT = "E7E7FF"
_SUCCESS = "71DD37"
_SUCCESS_LIGHT = "E8FBE0"
_DANGER = "FF3E1D"
_DANGER_LIGHT = "FFE0DA"
_WARNING = "FFAB00"
_WARNING_LIGHT = "FFF3D6"
_INFO = "03C3EC"
_GRAY_50 = "F5F5F9"
_GRAY_200 = "ECEEF1"
_GRAY_700 = "4B465C"
_WHITE = "FFFFFF"
_DARK = "2B2C40"

# ---------------------------------------------------------------------------
# Reusable styles
# ---------------------------------------------------------------------------
_THIN_BORDER = Border(
    left=Side(style="thin", color=_GRAY_200),
    right=Side(style="thin", color=_GRAY_200),
    top=Side(style="thin", color=_GRAY_200),
    bottom=Side(style="thin", color=_GRAY_200),
)

_HEADER_FONT = Font(name="Aptos", bold=True, color=_WHITE, size=11)
_HEADER_FILL = PatternFill(start_color=_PRIMARY, end_color=_PRIMARY, fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_BODY_FONT = Font(name="Aptos", size=11, color=_GRAY_700)
_BODY_ALT_FILL = PatternFill(start_color=_GRAY_50, end_color=_GRAY_50, fill_type="solid")

_TITLE_FONT = Font(name="Aptos", bold=True, size=24, color=_PRIMARY)
_SUBTITLE_FONT = Font(name="Aptos", size=13, color=_GRAY_700)
_SECTION_FONT = Font(name="Aptos", bold=True, size=14, color=_DARK)

_INCOME_FILL = PatternFill(start_color=_SUCCESS_LIGHT, end_color=_SUCCESS_LIGHT, fill_type="solid")
_EXPENSE_FILL = PatternFill(start_color=_DANGER_LIGHT, end_color=_DANGER_LIGHT, fill_type="solid")
_INCOME_FONT = Font(name="Aptos", size=11, color="28A745")
_EXPENSE_FONT = Font(name="Aptos", size=11, color=_DANGER)

_KPI_VALUE_FONT = Font(name="Aptos", bold=True, size=18, color=_DARK)
_KPI_LABEL_FONT = Font(name="Aptos", size=10, color=_GRAY_700)

_SUMMARY_LABEL_FONT = Font(name="Aptos", bold=True, size=11, color=_GRAY_700)
_SUMMARY_FILL = PatternFill(start_color=_PRIMARY_LIGHT, end_color=_PRIMARY_LIGHT, fill_type="solid")


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _write_header_row(ws, row: int, headers: list[str], start_col: int = 1):
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _write_data_row(ws, row: int, values: list, start_col: int = 1, alt: bool = False):
    """Write a data row with optional alternating background."""
    for col_idx, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if alt:
            cell.fill = _BODY_ALT_FILL


def _write_kpi(ws, row: int, col: int, label: str, value, fmt: str | None = None):
    """Write a KPI block (label + value) in two rows."""
    val_cell = ws.cell(row=row, column=col, value=value)
    val_cell.font = _KPI_VALUE_FONT
    val_cell.alignment = Alignment(horizontal="center")
    if fmt:
        val_cell.number_format = fmt

    lbl_cell = ws.cell(row=row + 1, column=col, value=label)
    lbl_cell.font = _KPI_LABEL_FONT
    lbl_cell.alignment = Alignment(horizontal="center")


def _write_summary_row(ws, row: int, label: str, value, col_label: int = 1,
                       col_value: int = 2, label_font=None, value_font=None,
                       num_fmt: str = "#,##0.00"):
    """Write a single summary label+value row with background highlight."""
    lbl = ws.cell(row=row, column=col_label, value=label)
    lbl.font = label_font or _SUMMARY_LABEL_FONT
    lbl.fill = _SUMMARY_FILL
    lbl.border = _THIN_BORDER

    val = ws.cell(row=row, column=col_value, value=value)
    val.font = value_font or _KPI_VALUE_FONT
    val.fill = _SUMMARY_FILL
    val.border = _THIN_BORDER
    if num_fmt:
        val.number_format = num_fmt

    # Fill remaining cells in the row up to col_value so it looks like a band
    for c in range(col_label + 1, col_value):
        cell = ws.cell(row=row, column=c)
        cell.fill = _SUMMARY_FILL
        cell.border = _THIN_BORDER


def _style_direction_cell(cell, direction: str):
    """Apply income/expense colour coding to a cell."""
    if direction == "in":
        cell.font = _INCOME_FONT
        cell.fill = _INCOME_FILL
    else:
        cell.font = _EXPENSE_FONT
        cell.fill = _EXPENSE_FILL


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_overview_sheet(wb: Workbook, session: Session, sections: list[str]):
    """Cover sheet with branding and summary KPIs."""
    ws = wb.active
    ws.title = _("dashboard")
    ws.sheet_properties.tabColor = _PRIMARY

    # Title
    ws.merge_cells("B2:G2")
    title_cell = ws["B2"]
    title_cell.value = "Yfine"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B3:G3")
    sub_cell = ws["B3"]
    sub_cell.value = _("export_excel_subtitle").replace("{date}", date.today().isoformat())
    sub_cell.font = _SUBTITLE_FONT

    # KPIs row
    row = 5
    ws.cell(row=row, column=2, value=_("dashboard")).font = _SECTION_FONT

    row = 7

    # Net worth
    sources = session.exec(select(Source)).all()
    net_worth_by_currency: dict[str, float] = {}
    for src in sources:
        in_sum = session.exec(
            select(func.coalesce(func.sum(Movement.amount), 0)).where(
                Movement.source_id == src.id, Movement.direction == "in"
            )
        ).one()
        out_sum = session.exec(
            select(func.coalesce(func.sum(Movement.amount), 0)).where(
                Movement.source_id == src.id, Movement.direction == "out"
            )
        ).one()
        bal = src.starting_balance + float(in_sum) - float(out_sum)
        net_worth_by_currency.setdefault(src.currency, 0.0)
        net_worth_by_currency[src.currency] += bal

    col = 2
    for curr, total in sorted(net_worth_by_currency.items()):
        _write_kpi(ws, row, col, f"{_('net_worth')} ({curr})", total, "#,##0.00")
        col += 1

    source_count = len(sources)
    _write_kpi(ws, row, col, _("total_sources"), source_count)
    col += 1

    mov_count = session.exec(select(func.count(Movement.id))).one()
    _write_kpi(ws, row, col, _("movements"), mov_count)
    col += 1

    tag_count = session.exec(select(func.count(Tag.id))).one()
    _write_kpi(ws, row, col, _("tags"), tag_count)

    # Sections included
    row = 10
    ws.cell(row=row, column=2, value=_("export_excel_contents")).font = _SECTION_FONT
    for i, section in enumerate(sections):
        cell = ws.cell(row=row + 1 + i, column=2, value=f"  {_(section)}")
        cell.font = _BODY_FONT

    # Monthly breakdown chart data (if movements selected)
    if "movements" in sections:
        _add_monthly_chart(ws, session, start_row=row + len(sections) + 3)

    _auto_width(ws, min_width=14)
    ws.column_dimensions["A"].width = 3  # left margin


def _add_monthly_chart(ws, session: Session, start_row: int):
    """Add an income vs expense bar chart by month."""
    today = date.today()
    months_data = []
    for i in range(11, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        month_start = date(y, m, 1)
        if m == 12:
            month_end = date(y + 1, 1, 1)
        else:
            month_end = date(y, m + 1, 1)

        inc = session.exec(
            select(func.coalesce(func.sum(Movement.amount), 0)).where(
                Movement.date >= month_start,
                Movement.date < month_end,
                Movement.direction == "in",
            )
        ).one()
        exp = session.exec(
            select(func.coalesce(func.sum(Movement.amount), 0)).where(
                Movement.date >= month_start,
                Movement.date < month_end,
                Movement.direction == "out",
            )
        ).one()
        label = month_start.strftime("%b %Y")
        months_data.append((label, float(inc), float(exp)))

    if not any(r[1] or r[2] for r in months_data):
        return

    ws.cell(row=start_row, column=2, value=_("export_excel_monthly_chart")).font = _SECTION_FONT

    # Write data for chart
    data_row = start_row + 1
    ws.cell(row=data_row, column=2, value=_("date"))
    ws.cell(row=data_row, column=3, value=_("income"))
    ws.cell(row=data_row, column=4, value=_("expense"))
    for i, (label, inc, exp) in enumerate(months_data):
        ws.cell(row=data_row + 1 + i, column=2, value=label)
        ws.cell(row=data_row + 1 + i, column=3, value=inc)
        ws.cell(row=data_row + 1 + i, column=4, value=exp)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.y_axis.title = _("amount")
    chart.x_axis.title = _("date")
    chart.width = 28
    chart.height = 14

    data_ref = Reference(ws, min_col=3, max_col=4, min_row=data_row, max_row=data_row + 12)
    cats = Reference(ws, min_col=2, min_row=data_row + 1, max_row=data_row + 12)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = _SUCCESS
    chart.series[1].graphicalProperties.solidFill = _DANGER

    ws.add_chart(chart, f"B{data_row + 14}")


def _build_sources_sheet(wb: Workbook, session: Session):
    """Sources sheet with totals at top, then data table."""
    ws = wb.create_sheet(_("sources"))
    ws.sheet_properties.tabColor = _INFO

    sources = session.exec(select(Source)).all()

    # Pre-compute balances
    balances = []
    net_by_currency: dict[str, float] = {}
    for src in sources:
        in_sum = session.exec(
            select(func.coalesce(func.sum(Movement.amount), 0)).where(
                Movement.source_id == src.id, Movement.direction == "in"
            )
        ).one()
        out_sum = session.exec(
            select(func.coalesce(func.sum(Movement.amount), 0)).where(
                Movement.source_id == src.id, Movement.direction == "out"
            )
        ).one()
        balance = src.starting_balance + float(in_sum) - float(out_sum)
        balances.append(balance)
        net_by_currency.setdefault(src.currency, 0.0)
        net_by_currency[src.currency] += balance

    # --- Summary at top ---
    row = 1
    ws.cell(row=row, column=1, value=_("net_worth")).font = _SECTION_FONT
    row = 2
    for curr, total in sorted(net_by_currency.items()):
        _write_summary_row(ws, row, f"{_('net_worth')} ({curr})", total,
                           value_font=Font(name="Aptos", bold=True, size=14,
                                           color="28A745" if total >= 0 else _DANGER))
        row += 1

    # Blank separator
    row += 1

    # --- Data table ---
    headers = [_("name"), _("currency"), _("starting_balance"), _("current_balance"),
               _("exclude_from_stats")]
    table_start = row
    _write_header_row(ws, table_start, headers)

    for i, src in enumerate(sources):
        r = table_start + 1 + i
        balance = balances[i]
        values = [src.name, src.currency, src.starting_balance, balance,
                  _("exclude_from_stats") if src.exclude_from_stats else ""]
        _write_data_row(ws, r, values, alt=i % 2 == 1)

        ws.cell(row=r, column=3).number_format = "#,##0.00"
        bal_cell = ws.cell(row=r, column=4)
        bal_cell.number_format = "#,##0.00"
        bal_cell.font = _INCOME_FONT if balance >= 0 else _EXPENSE_FONT

    _auto_width(ws)
    ws.freeze_panes = f"A{table_start + 1}"


def _build_movements_sheet(wb: Workbook, session: Session):
    """Movements sheet — summary on top, data table below."""
    ws = wb.create_sheet(_("movements"))
    ws.sheet_properties.tabColor = _SUCCESS

    # Build lookups
    sources_map = {s.id: s.name for s in session.exec(select(Source)).all()}
    tags_map = {t.id: t.name for t in session.exec(select(Tag)).all()}
    mov_tags: dict[int, list[str]] = {}
    for mt in session.exec(select(MovementTag)).all():
        mov_tags.setdefault(mt.movement_id, []).append(tags_map.get(mt.tag_id, ""))

    movements = session.exec(select(Movement).order_by(Movement.date.desc())).all()

    # Pre-compute totals
    total_in = sum(m.amount for m in movements if m.direction == "in")
    total_out = sum(m.amount for m in movements if m.direction == "out")
    net = total_in - total_out

    # --- Summary at top ---
    row = 1
    _write_summary_row(ws, row, _("income"), total_in,
                       label_font=Font(name="Aptos", bold=True, size=11, color="28A745"),
                       value_font=Font(name="Aptos", bold=True, size=14, color="28A745"))
    row += 1
    _write_summary_row(ws, row, _("expense"), total_out,
                       label_font=Font(name="Aptos", bold=True, size=11, color=_DANGER),
                       value_font=Font(name="Aptos", bold=True, size=14, color=_DANGER))
    row += 1
    _write_summary_row(ws, row, "Net", net,
                       value_font=Font(name="Aptos", bold=True, size=16, color=_DARK))
    row += 1
    _write_summary_row(ws, row, _("movements"), len(movements),
                       value_font=Font(name="Aptos", bold=True, size=12, color=_GRAY_700),
                       num_fmt="#,##0")

    # Blank separator
    row += 2

    # --- Data table ---
    headers = [_("date"), _("source"), _("direction"), _("amount"), _("note"), _("tags"),
               _("transfer")]
    table_start = row
    _write_header_row(ws, table_start, headers)

    for i, mov in enumerate(movements):
        r = table_start + 1 + i
        source_name = sources_map.get(mov.source_id, _("external"))
        direction_label = _("income") if mov.direction == "in" else _("expense")
        tag_str = ", ".join(mov_tags.get(mov.id, []))
        transfer_label = _("transfer") if mov.transfer_pair_id else ""

        values = [mov.date, source_name, direction_label, mov.amount, mov.note or "",
                  tag_str, transfer_label]
        _write_data_row(ws, r, values, alt=i % 2 == 1)

        ws.cell(row=r, column=1).number_format = "YYYY-MM-DD"
        _style_direction_cell(ws.cell(row=r, column=3), mov.direction)
        amt_cell = ws.cell(row=r, column=4)
        amt_cell.number_format = "#,##0.00"
        amt_cell.font = _INCOME_FONT if mov.direction == "in" else _EXPENSE_FONT

    _auto_width(ws)
    ws.freeze_panes = f"A{table_start + 1}"
    if movements:
        ws.auto_filter.ref = f"A{table_start}:G{table_start + len(movements)}"


def _build_tags_sheet(wb: Workbook, session: Session):
    """Tags sheet with chart on the right, data table on the left."""
    ws = wb.create_sheet(_("tags"))
    ws.sheet_properties.tabColor = _WARNING

    tags = session.exec(select(Tag)).all()

    # Pre-compute counts
    tag_data = []
    for tag in tags:
        mov_count = session.exec(
            select(func.count(MovementTag.movement_id)).where(MovementTag.tag_id == tag.id)
        ).one()
        sav_count = session.exec(
            select(func.count(SavingTag.saving_id)).where(SavingTag.tag_id == tag.id)
        ).one()
        tag_data.append((tag, mov_count, sav_count))

    total_usage = sum(mc for _t, mc, _s in tag_data)

    # --- Summary at top ---
    row = 1
    _write_summary_row(ws, row, _("tags"), len(tags),
                       value_font=Font(name="Aptos", bold=True, size=14, color=_DARK),
                       num_fmt="#,##0")
    row += 1
    _write_summary_row(ws, row, _("export_excel_usage_movements"), total_usage,
                       value_font=Font(name="Aptos", bold=True, size=14, color=_DARK),
                       num_fmt="#,##0")

    # Blank separator
    row += 2

    # --- Data table ---
    headers = [_("name"), _("tag_color"), _("export_excel_usage_movements"),
               _("export_excel_usage_savings")]
    table_start = row
    _write_header_row(ws, table_start, headers)

    for i, (tag, mov_count, sav_count) in enumerate(tag_data):
        r = table_start + 1 + i
        values = [tag.name, tag.color or "", mov_count, sav_count]
        _write_data_row(ws, r, values, alt=i % 2 == 1)

        if tag.color:
            color_cell = ws.cell(row=r, column=2)
            hex_color = tag.color.lstrip("#").upper()
            if len(hex_color) == 6:
                color_cell.fill = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid")
                color_cell.font = Font(name="Aptos", size=11, color=hex_color)

    # Pie chart — placed to the right of the table
    if tag_data:
        _add_tag_pie_chart(ws, tag_data, table_start)

    _auto_width(ws)


def _add_tag_pie_chart(ws, tag_data: list, table_start: int):
    """Pie chart showing distribution of tag usage across movements."""
    # Write chart source data in columns F-G (hidden area to the right)
    data_row = table_start
    ws.cell(row=data_row, column=6, value=_("name"))
    ws.cell(row=data_row, column=7, value=_("export_excel_usage_movements"))

    has_data = False
    for i, (tag, mov_count, _sc) in enumerate(tag_data):
        ws.cell(row=data_row + 1 + i, column=6, value=tag.name)
        ws.cell(row=data_row + 1 + i, column=7, value=mov_count)
        if mov_count > 0:
            has_data = True

    if not has_data:
        return

    chart = PieChart()
    chart.style = 10
    chart.width = 18
    chart.height = 14

    data_ref = Reference(ws, min_col=7, min_row=data_row, max_row=data_row + len(tag_data))
    cats = Reference(ws, min_col=6, min_row=data_row + 1, max_row=data_row + len(tag_data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True

    ws.add_chart(chart, f"F1")


def _build_recurring_sheet(wb: Workbook, session: Session):
    """Recurring items sheet — monthly equivalence summary on top."""
    ws = wb.create_sheet(_("recurring"))
    ws.sheet_properties.tabColor = _DANGER

    sources_map = {s.id: s.name for s in session.exec(select(Source)).all()}
    items = session.exec(select(RecurringItem).order_by(RecurringItem.next_due_date)).all()

    freq_monthly_factor = {"daily": 30, "weekly": 4.33, "monthly": 1, "yearly": 1 / 12}

    # Pre-compute monthly equivalence
    total_monthly_in = 0.0
    total_monthly_out = 0.0
    for item in items:
        factor = freq_monthly_factor.get(item.frequency, 1)
        monthly = item.amount * factor
        if item.direction == "in":
            total_monthly_in += monthly
        else:
            total_monthly_out += monthly

    monthly_net = total_monthly_in - total_monthly_out

    # --- Summary at top ---
    row = 1
    ws.cell(row=row, column=1, value=_("export_excel_monthly_equiv")).font = _SECTION_FONT
    row = 2
    _write_summary_row(ws, row, _("income"), round(total_monthly_in, 2),
                       label_font=Font(name="Aptos", bold=True, size=11, color="28A745"),
                       value_font=Font(name="Aptos", bold=True, size=14, color="28A745"))
    row += 1
    _write_summary_row(ws, row, _("expense"), round(total_monthly_out, 2),
                       label_font=Font(name="Aptos", bold=True, size=11, color=_DANGER),
                       value_font=Font(name="Aptos", bold=True, size=14, color=_DANGER))
    row += 1
    _write_summary_row(ws, row, "Net", round(monthly_net, 2),
                       value_font=Font(name="Aptos", bold=True, size=16, color=_DARK))
    row += 1
    _write_summary_row(ws, row, _("recurring"), len(items),
                       value_font=Font(name="Aptos", bold=True, size=12, color=_GRAY_700),
                       num_fmt="#,##0")

    # Blank separator
    row += 2

    # --- Data table ---
    headers = [_("name"), _("amount"), _("direction"), _("currency"), _("frequency"),
               _("source"), _("apply_mode"), _("next_due_date"), _("start_date"), _("end_date")]
    table_start = row
    _write_header_row(ws, table_start, headers)

    for i, item in enumerate(items):
        r = table_start + 1 + i
        source_name = sources_map.get(item.source_id, "—")
        direction_label = _("income") if item.direction == "in" else _("expense")
        freq_label = _(item.frequency)
        mode_label = _("automatic") if item.apply_mode == "auto" else _("manual_confirm")

        values = [item.name, item.amount, direction_label, item.currency, freq_label,
                  source_name, mode_label, item.next_due_date, item.start_date,
                  item.end_date or ""]
        _write_data_row(ws, r, values, alt=i % 2 == 1)

        ws.cell(row=r, column=2).number_format = "#,##0.00"
        _style_direction_cell(ws.cell(row=r, column=3), item.direction)
        ws.cell(row=r, column=8).number_format = "YYYY-MM-DD"
        ws.cell(row=r, column=9).number_format = "YYYY-MM-DD"

    _auto_width(ws)
    ws.freeze_panes = f"A{table_start + 1}"


def _build_savings_sheet(wb: Workbook, session: Session):
    """Savings sheet — total at top, data below."""
    ws = wb.create_sheet(_("savings"))
    ws.sheet_properties.tabColor = "4CAF50"

    tags_map = {t.id: t.name for t in session.exec(select(Tag)).all()}
    sav_tags: dict[int, list[str]] = {}
    for st in session.exec(select(SavingTag)).all():
        sav_tags.setdefault(st.saving_id, []).append(tags_map.get(st.tag_id, ""))

    savings = session.exec(select(Saving).order_by(Saving.date.desc())).all()

    total = sum(s.amount for s in savings)

    # --- Summary at top ---
    row = 1
    _write_summary_row(ws, row, _("savings_total"), total,
                       value_font=Font(name="Aptos", bold=True, size=16, color="28A745"))
    row += 1
    _write_summary_row(ws, row, _("savings"), len(savings),
                       value_font=Font(name="Aptos", bold=True, size=12, color=_GRAY_700),
                       num_fmt="#,##0")

    # Blank separator
    row += 2

    # --- Data table ---
    headers = [_("date"), _("amount"), _("currency"), _("saving_description"), _("note"),
               _("tags")]
    table_start = row
    _write_header_row(ws, table_start, headers)

    for i, sav in enumerate(savings):
        r = table_start + 1 + i
        tag_str = ", ".join(sav_tags.get(sav.id, []))
        values = [sav.date, sav.amount, sav.currency, sav.description or "", sav.note or "",
                  tag_str]
        _write_data_row(ws, r, values, alt=i % 2 == 1)
        ws.cell(row=r, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=r, column=2).number_format = "#,##0.00"
        ws.cell(row=r, column=2).font = _INCOME_FONT

    _auto_width(ws)
    ws.freeze_panes = f"A{table_start + 1}"


def _build_whims_sheet(wb: Workbook, session: Session):
    """Whims sheet — totals by status at top, data below."""
    ws = wb.create_sheet(_("whims"))
    ws.sheet_properties.tabColor = _WARNING

    sources_map = {s.id: s.name for s in session.exec(select(Source)).all()}
    whims = session.exec(select(Whim).order_by(Whim.status, Whim.priority.desc())).all()

    pending_total = sum(w.amount for w in whims if w.status == "pending")
    purchased_total = sum(w.amount for w in whims if w.status == "purchased")
    dismissed_count = sum(1 for w in whims if w.status == "dismissed")

    status_labels = {
        "pending": _("whim_pending"),
        "purchased": _("whim_purchased"),
        "dismissed": _("whim_dismissed"),
    }
    priority_labels = {
        "low": _("priority_low"),
        "medium": _("priority_medium"),
        "high": _("priority_high"),
    }

    # --- Summary at top ---
    row = 1
    _write_summary_row(ws, row, _("whim_pending"), pending_total,
                       label_font=Font(name="Aptos", bold=True, size=11, color=_WARNING),
                       value_font=Font(name="Aptos", bold=True, size=16, color=_WARNING))
    row += 1
    _write_summary_row(ws, row, _("whim_purchased"), purchased_total,
                       label_font=Font(name="Aptos", bold=True, size=11, color="28A745"),
                       value_font=Font(name="Aptos", bold=True, size=14, color="28A745"))
    row += 1
    _write_summary_row(ws, row, _("whim_dismissed"), dismissed_count,
                       label_font=Font(name="Aptos", bold=True, size=11, color=_GRAY_700),
                       value_font=Font(name="Aptos", bold=True, size=12, color=_GRAY_700),
                       num_fmt="#,##0")
    row += 1
    _write_summary_row(ws, row, _("whims"), len(whims),
                       value_font=Font(name="Aptos", bold=True, size=12, color=_GRAY_700),
                       num_fmt="#,##0")

    # Blank separator
    row += 2

    # --- Data table ---
    headers = [_("name"), _("amount"), _("currency"), _("whim_priority"), _("source"),
               _("export_excel_status"), _("note"), _("whim_link")]
    table_start = row
    _write_header_row(ws, table_start, headers)

    for i, whim in enumerate(whims):
        r = table_start + 1 + i
        source_name = sources_map.get(whim.source_id, "—")
        values = [
            whim.name, whim.amount, whim.currency,
            priority_labels.get(whim.priority, whim.priority),
            source_name, status_labels.get(whim.status, whim.status),
            whim.note or "", whim.url or "",
        ]
        _write_data_row(ws, r, values, alt=i % 2 == 1)

        ws.cell(row=r, column=2).number_format = "#,##0.00"

        # Status colouring
        status_cell = ws.cell(row=r, column=6)
        if whim.status == "purchased":
            status_cell.font = _INCOME_FONT
            status_cell.fill = _INCOME_FILL
        elif whim.status == "dismissed":
            status_cell.font = Font(name="Aptos", size=11, color=_GRAY_700)
        else:
            status_cell.font = Font(name="Aptos", size=11, color=_WARNING)

        # Priority colouring
        prio_cell = ws.cell(row=r, column=4)
        if whim.priority == "high":
            prio_cell.font = _EXPENSE_FONT
        elif whim.priority == "medium":
            prio_cell.font = Font(name="Aptos", size=11, color=_WARNING)

    _auto_width(ws)
    ws.freeze_panes = f"A{table_start + 1}"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

# Allowed section keys
EXPORTABLE_SECTIONS = ["sources", "movements", "tags", "recurring", "savings", "whims"]

_BUILDERS = {
    "sources": _build_sources_sheet,
    "movements": _build_movements_sheet,
    "tags": _build_tags_sheet,
    "recurring": _build_recurring_sheet,
    "savings": _build_savings_sheet,
    "whims": _build_whims_sheet,
}


def export_excel(session: Session, sections: list[str]) -> bytes:
    """Generate a styled Excel workbook with the selected sections.

    Args:
        session: DB session.
        sections: list of section keys from EXPORTABLE_SECTIONS.

    Returns:
        bytes of the .xlsx file.
    """
    valid = [s for s in sections if s in _BUILDERS]
    if not valid:
        valid = list(_BUILDERS.keys())

    wb = Workbook()

    # Always build overview
    _build_overview_sheet(wb, session, valid)

    # Build requested sheets
    for section in valid:
        _BUILDERS[section](wb, session)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
