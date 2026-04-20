"""XLSX parser via openpyxl — reuses CSV column-mapping heuristics."""
import hashlib
import io
from datetime import date, datetime

from services.importers.base import ParsedMovement, ParseResult
from services.importers.csv_parser import _guess_column_map, _parse_amount, _try_parse_date


class XlsxParser:
    format_key = "xlsx"
    display_name = "Excel (XLSX)"
    file_extensions = (".xlsx",)

    def sniff(self, raw: bytes) -> bool:
        # XLSX files are ZIP archives starting with PK\x03\x04
        if not raw.startswith(b"PK\x03\x04"):
            return False
        # Quick check for xlsx-specific marker without full unzip
        return b"xl/workbook.xml" in raw[:4096] or b"[Content_Types].xml" in raw[:4096]

    def parse(self, raw: bytes, options: dict | None = None) -> ParseResult:
        try:
            import openpyxl
        except ImportError:
            return ParseResult(movements=[], warnings=["openpyxl_not_installed"])

        opts = options or {}
        column_map = opts.get("column_map")
        decimal_separator = opts.get("decimal_separator", ".")
        date_format = opts.get("date_format")
        sheet_name = opts.get("sheet_name")
        header_row = int(opts.get("header_row", 1) or 1)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            return ParseResult(movements=[], warnings=[f"parse_error:{exc.__class__.__name__}"])

        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        if ws is None:
            return ParseResult(movements=[], warnings=["no_sheets"])

        rows_iter = ws.iter_rows(values_only=True)
        all_rows: list[tuple] = []
        for r in rows_iter:
            all_rows.append(r)
            if len(all_rows) > 100000:
                break

        if len(all_rows) < header_row:
            return ParseResult(movements=[], warnings=["empty_file"])

        headers_row = all_rows[header_row - 1]
        headers = [str(h).strip() if h is not None else "" for h in headers_row]
        data_rows = all_rows[header_row:]

        if not column_map:
            column_map = _guess_column_map(headers)
        if not column_map:
            return ParseResult(
                movements=[],
                warnings=[f"needs_mapping:{','.join(headers)}"],
            )

        header_index: dict[str, int] = {}
        for fld, header_name in column_map.items():
            if header_name in headers:
                header_index[fld] = headers.index(header_name)

        movements: list[ParsedMovement] = []
        detected_currency: str | None = None
        warnings: list[str] = []

        for row_num, row in enumerate(data_rows, start=header_row + 1):
            if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue

            row_key = "|".join(str(c) if c is not None else "" for c in row).encode("utf-8")
            raw_hash = hashlib.sha256(row_key).hexdigest()

            d_idx = header_index.get("date")
            d_val = row[d_idx] if d_idx is not None and d_idx < len(row) else None
            if isinstance(d_val, (date, datetime)):
                parsed_date = d_val.date() if isinstance(d_val, datetime) else d_val
            elif isinstance(d_val, str):
                parsed_date = _try_parse_date(d_val, date_format)
            else:
                parsed_date = None
            if parsed_date is None:
                warnings.append(f"row_{row_num}_bad_date")
                continue

            amount: float | None = None
            direction: str | None = None

            if "amount_in" in header_index and "amount_out" in header_index:
                in_val = row[header_index["amount_in"]] if header_index["amount_in"] < len(row) else None
                out_val = row[header_index["amount_out"]] if header_index["amount_out"] < len(row) else None
                in_f = _normalize_number(in_val, decimal_separator)
                out_f = _normalize_number(out_val, decimal_separator)
                if in_f and in_f > 0:
                    amount = in_f
                    direction = "in"
                elif out_f and out_f > 0:
                    amount = out_f
                    direction = "out"
            else:
                a_idx = header_index.get("amount")
                if a_idx is not None and a_idx < len(row):
                    a_val = row[a_idx]
                    a_f = _normalize_number(a_val, decimal_separator)
                    if a_f is None:
                        warnings.append(f"row_{row_num}_bad_amount")
                        continue
                    direction = "in" if a_f >= 0 else "out"
                    amount = abs(a_f)

            if amount is None or amount == 0 or direction not in ("in", "out"):
                warnings.append(f"row_{row_num}_zero_or_invalid")
                continue

            note = None
            n_idx = header_index.get("note")
            if n_idx is not None and n_idx < len(row):
                n_val = row[n_idx]
                if n_val is not None:
                    note = str(n_val).strip() or None

            currency = None
            c_idx = header_index.get("currency")
            if c_idx is not None and c_idx < len(row):
                c_val = row[c_idx]
                if c_val is not None:
                    currency = str(c_val).strip().upper() or None
                    if currency and not detected_currency:
                        detected_currency = currency

            movements.append(ParsedMovement(
                date=parsed_date,
                amount=round(amount, 2),
                direction=direction,
                note=note,
                raw_hash=raw_hash,
                currency=currency,
            ))

        return ParseResult(
            movements=movements,
            detected_currency=detected_currency,
            detected_source_hint=None,
            warnings=warnings,
        )


def _normalize_number(value, decimal_separator: str = ".") -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return _parse_amount(str(value), decimal_separator)
